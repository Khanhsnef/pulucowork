import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

# Create Workbook
wb = openpyxl.Workbook()
wb.remove(wb.active) # Remove default sheet

# Color Palette (Sleek Modern Corporate Theme)
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")

SECTION_DM_FILL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid") # Teal for DM
SECTION_QM_FILL = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid") # Indigo for QM

TITLE_FONT = Font(name="Arial", size=14, bold=True, color="0F172A")
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="475569")
BOLD_FONT = Font(name="Arial", size=10, bold=True)
REGULAR_FONT = Font(name="Arial", size=10)

# Status Fills
STATUS_APPROVED = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Light Green
STATUS_REJECTED = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red
STATUS_PENDING = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light Yellow
STATUS_PROCESSING = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid") # Light Blue

THIN_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

# -------------------------------------------------------------
# TAB 1: 01_ALL_TAG_REQUESTS (Master Tracker)
# -------------------------------------------------------------
ws1 = wb.create_sheet(title="01_ALL_TAG_REQUESTS")
ws1.views.sheetView[0].showGridLines = True

# Title Block
ws1.merge_cells("A1:U1")
ws1["A1"] = "HỆ THỐNG QUẢN LÝ & ĐỀ XUẤT TAG TÀI XẾ - MASTER REQUEST TRACKER"
ws1["A1"].font = TITLE_FONT
ws1["A1"].alignment = Alignment(vertical="center")

ws1.merge_cells("A2:U2")
ws1["A2"] = "Quy trình 3 bước: Requesting Team Submission -> DM Review & Enrich -> QM Tag Execution"
ws1["A2"].font = SUBTITLE_FONT

headers_ws1 = [
    # Request Info
    "Mã Request", "Ngày Tạo", "Team Đề Xuất", "Người Đề Xuất", "Loại Tag", "Tên Tag Yêu Cầu", 
    "Số Lượng TX", "Lý Do Kinh Doanh / Mục Tiêu", "Thời Gian Áp Dụng",
    # DM Review Block
    "DM Reviewer", "Ngày DM Review", "DM Quyết Định", "Lý Do Tuần Thủ / Từ Chối", "DM Tag Code Chuẩn", "Trạng Thái Chuyển QM",
    # QM Execution Block
    "QM Specialist", "Ngày QM Nhận", "QM Trạng Thái Add Tag", "Ngày Add Tag Xong", "Số TX Add Thành Công", "Ghi Chú QM / Batch ID"
]

# Write Headers
ws1.row_dimensions[4].height = 28
for col_num, header_title in enumerate(headers_ws1, 1):
    cell = ws1.cell(row=4, column=col_num)
    cell.value = header_title
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Specific section header colors
    if col_num <= 9:
        cell.fill = HEADER_FILL
    elif col_num <= 15:
        cell.fill = SECTION_DM_FILL
    else:
        cell.fill = SECTION_QM_FILL

# Sample Data Rows
sample_data_ws1 = [
    [
        "REQ-2026-001", "2026-07-25 09:00", "Business Ops", "Nguyễn Văn A", "Priority Dispatch", "TAG_HUB_LH_OFFLOAD_VIP",
        150, "Ưu tiên phát đơn liên vùng tuyến SOC Củ Chi", "2026-08-01 đến 2026-08-31",
        "Trần Lead DM", "2026-07-25 11:30", "APPROVED", "Đạt chuẩn SLA & Không có driver vi phạm fraud", "DM_TAG_HUB_LH_VIP_V1", "READY_FOR_QM",
        "Lê QM Spec", "2026-07-25 13:00", "TAGGED_SUCCESS", "2026-07-25 14:15", 150, "BATCH_QM_88412 - Succ 100%"
    ],
    [
        "REQ-2026-002", "2026-07-26 10:15", "Marketing", "Phạm Thị B", "Incentive Campaign", "TAG_MAR_PEAK_BONUS_20K",
        450, "Thưởng nóng tài xế chạy ca tối 18h-22h", "2026-07-28 đến 2026-07-30",
        "Trần Lead DM", "2026-07-26 14:00", "APPROVED", "Đã chốt ngân sách Marketing với BI", "DM_TAG_BONUS_NIGHT_20K", "READY_FOR_QM",
        "Lê QM Spec", "2026-07-26 15:30", "TAGGED_SUCCESS", "2026-07-26 16:45", 448, "BATCH_QM_88425 - 2 TX sai số ĐT"
    ],
    [
        "REQ-2026-003", "2026-07-27 14:20", "Hub Operations", "Hoàng Văn C", "Area Restriction", "TAG_ZONE_RESTRICT_TAN_BINH",
        85, "Chỉ cho phép tài xế có đào tạo Hub Tân Bình vội đơn", "2026-08-01 đến 2026-10-31",
        "Trần Lead DM", "2026-07-27 16:00", "REJECTED", "Danh sách TX thiếu bằng chứng qua lớp đào tạo", "N/A", "CANCELLED",
        "-", "-", "N/A", "-", 0, "Yêu cầu bị DM từ chối"
    ],
    [
        "REQ-2026-004", "2026-07-28 08:45", "Customer Service", "Đỗ Thị D", "Special Training", "TAG_CS_VIP_HANDLING",
        30, "Tài xế chuyên xử lý đơn VIP hàng giá trị cao", "Dài hạn",
        "Trần Lead DM", "2026-07-28 10:00", "APPROVED", "CS đã duyệt danh sách 30 tài xế chuẩn 5 sao", "DM_TAG_CS_VIP_5STAR", "READY_FOR_QM",
        "Vũ QM Spec", "2026-07-28 10:30", "PROCESSING", "-", 0, "Đang kiểm tra tool add tag hệ thống"
    ],
    [
        "REQ-2026-005", "2026-07-29 09:30", "Risk Management", "Vũ Văn E", "Penalty / Block", "TAG_RISK_SUSPEND_HUB",
        12, "Tài xế dấu hiệu gian lận Hub offload", "2026-07-29 đến khi có thông báo",
        "Trần Lead DM", "2026-07-29 10:00", "PENDING", "Đang chờ BI kiểm tra logs GPS", "-", "HOLD",
        "-", "-", "PENDING_QM", "-", 0, "Chờ DM duyệt xong"
    ]
]

for row_idx, row_data in enumerate(sample_data_ws1, 5):
    ws1.row_dimensions[row_idx].height = 22
    for col_idx, val in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font = REGULAR_FONT
        cell.border = THIN_BORDER
        
        # Alignment
        if col_idx in [1, 2, 10, 11, 12, 15, 16, 17, 18, 19]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx in [7, 20]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
        # Highlight Status
        if col_idx == 12: # DM Decision
            if val == "APPROVED":
                cell.fill = STATUS_APPROVED
                cell.font = Font(name="Arial", size=10, bold=True, color="166534")
            elif val == "REJECTED":
                cell.fill = STATUS_REJECTED
                cell.font = Font(name="Arial", size=10, bold=True, color="991B1B")
            elif val == "PENDING":
                cell.fill = STATUS_PENDING
                cell.font = Font(name="Arial", size=10, bold=True, color="92400E")
                
        if col_idx == 18: # QM Status
            if val == "TAGGED_SUCCESS":
                cell.fill = STATUS_APPROVED
                cell.font = Font(name="Arial", size=10, bold=True, color="166534")
            elif val == "PROCESSING":
                cell.fill = STATUS_PROCESSING
                cell.font = Font(name="Arial", size=10, bold=True, color="3730A3")
            elif val == "PENDING_QM":
                cell.fill = STATUS_PENDING
                cell.font = Font(name="Arial", size=10, bold=True, color="92400E")

# Data Validations for Tab 1
dv_dept = DataValidation(type="list", formula1='"Business Ops, Marketing, Hub Operations, Customer Service, Risk Management, Fleet Management"', allow_blank=True)
dv_tag_type = DataValidation(type="list", formula1='"Priority Dispatch, Incentive Campaign, Area Restriction, Special Training, Penalty / Block, Account Governance"', allow_blank=True)
dv_dm_decision = DataValidation(type="list", formula1='"APPROVED, REJECTED, PENDING, NEED_MORE_INFO"', allow_blank=True)
dv_qm_handover = DataValidation(type="list", formula1='"READY_FOR_QM, HOLD, CANCELLED"', allow_blank=True)
dv_qm_status = DataValidation(type="list", formula1='"PENDING_QM, PROCESSING, TAGGED_SUCCESS, PARTIAL_SUCCESS, FAILED, N/A"', allow_blank=True)

ws1.add_data_validation(dv_dept)
ws1.add_data_validation(dv_tag_type)
ws1.add_data_validation(dv_dm_decision)
ws1.add_data_validation(dv_qm_handover)
ws1.add_data_validation(dv_qm_status)

dv_dept.add("C5:C500")
dv_tag_type.add("E5:E500")
dv_dm_decision.add("L5:L500")
dv_qm_handover.add("O5:O500")
dv_qm_status.add("R5:R500")


# -------------------------------------------------------------
# TAB 2: 02_DM_LEAD_REVIEW_CONSOLE
# -------------------------------------------------------------
ws2 = wb.create_sheet(title="02_DM_LEAD_REVIEW_CONSOLE")
ws2.views.sheetView[0].showGridLines = True

ws2.merge_cells("A1:M1")
ws2["A1"] = "BÀN TRÒN KIỂM DUYỆT CỦA DM LEAD (DM REVIEW & ENRICHMENT CONSOLE)"
ws2["A1"].font = TITLE_FONT
ws2["A1"].alignment = Alignment(vertical="center")

ws2.merge_cells("A2:M2")
ws2["A2"] = "Dành riêng cho Lead DM để Thẩm định Yêu cầu, Kiểm tra Fraud/Policy & Gán Mã Tag Chuẩn trước khi gửi QM"
ws2["A2"].font = SUBTITLE_FONT

headers_ws2 = [
    "Mã Request", "Ngày Tạo", "Team Yêu Cầu", "Tên Tag Yêu Cầu", "Số Lượng TX", 
    "Quyết Định DM", "Lý Do / Tiêu Chuẩn Thẩm Định", "DM Tag Code Chuẩn", "Thời Hạn Tag",
    "Chuyển QM (Y/N)", "Thời Gian Phản Hồi DM (Giờ)", "SLA Review Check", "Ghi Chú DM"
]

ws2.row_dimensions[4].height = 28
for col_num, title in enumerate(headers_ws2, 1):
    cell = ws2.cell(row=4, column=col_num, value=title)
    cell.font = HEADER_FONT
    cell.fill = SECTION_DM_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

sample_data_ws2 = [
    ["REQ-2026-001", "2026-07-25 09:00", "Business Ops", "TAG_HUB_LH_OFFLOAD_VIP", 150, "APPROVED", "TX đạt 98% acceptance rate, không vi phạm", "DM_TAG_HUB_LH_VIP_V1", "30 ngày", "READY_FOR_QM", "=ROUND((INT(B5)-INT(B5))*24,1)", "DAT_SLA", "Đã đính kèm file driver ID"],
    ["REQ-2026-002", "2026-07-26 10:15", "Marketing", "TAG_MAR_PEAK_BONUS_20K", 450, "APPROVED", "Đã xác nhận với BI & Trưởng phòng MKT", "DM_TAG_BONUS_NIGHT_20K", "3 ngày", "READY_FOR_QM", "3.75", "DAT_SLA", "Tài trợ 100% budget Marketing"],
    ["REQ-2026-003", "2026-07-27 14:20", "Hub Operations", "TAG_ZONE_RESTRICT_TAN_BINH", 85, "REJECTED", "Thiếu danh sách chứng chỉ đào tạo Hub", "N/A", "N/A", "CANCELLED", "1.67", "DAT_SLA", "Yêu cầu Ops gửi lại kèm file PDF kết quả thi"],
    ["REQ-2026-004", "2026-07-28 08:45", "Customer Service", "TAG_CS_VIP_HANDLING", 30, "APPROVED", "Driver 5-star rating > 4.95", "DM_TAG_CS_VIP_5STAR", "90 ngày", "READY_FOR_QM", "1.25", "DAT_SLA", "Priority tag"],
    ["REQ-2026-005", "2026-07-29 09:30", "Risk Management", "TAG_RISK_SUSPEND_HUB", 12, "PENDING", "Đang chờ BI extract raw log trùng tọa độ GPS", "-", "14 ngày", "HOLD", "0.5", "DANG_XULY", "Urgent flag từ Risk Team"]
]

for row_idx, row_data in enumerate(sample_data_ws2, 5):
    ws2.row_dimensions[row_idx].height = 22
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = REGULAR_FONT
        cell.border = THIN_BORDER
        if col_idx in [1, 2, 6, 10, 12]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 5:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")


# -------------------------------------------------------------
# TAB 3: 03_QM_HANDOVER_QUEUE
# -------------------------------------------------------------
ws3 = wb.create_sheet(title="03_QM_HANDOVER_QUEUE")
ws3.views.sheetView[0].showGridLines = True

ws3.merge_cells("A1:K1")
ws3["A1"] = "DANH SÁCH CHỜ QM THỰC HIỆN ADD TAG (QM EXECUTION QUEUE)"
ws3["A1"].font = TITLE_FONT
ws3["A1"].alignment = Alignment(vertical="center")

ws3.merge_cells("A2:K2")
ws3["A2"] = "Chỉ hiển thị các Yêu cầu ĐÃ ĐƯỢC DM DUYỆT (DM Approved) để QM tiếp nhận và thao tác hệ thống"
ws3["A2"].font = SUBTITLE_FONT

headers_ws3 = [
    "Mã Request", "Ngày DM Chuyển", "Team Yêu Cầu", "DM Tag Code Chuẩn", "Số Lượng TX Cần Tag", 
    "QM Chuyên Viên", "Trạng Thái Add Tag", "Ngày Hoàn Thành", "Số TX Thành Công", "Tỷ Lệ Thành Công", "QM System Batch Ref / Note"
]

ws3.row_dimensions[4].height = 28
for col_num, title in enumerate(headers_ws3, 1):
    cell = ws3.cell(row=4, column=col_num, value=title)
    cell.font = HEADER_FONT
    cell.fill = SECTION_QM_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

sample_data_ws3 = [
    ["REQ-2026-001", "2026-07-25 11:30", "Business Ops", "DM_TAG_HUB_LH_VIP_V1", 150, "Lê QM Spec", "TAGGED_SUCCESS", "2026-07-25 14:15", 150, "=I5/E5", "BATCH_QM_88412 - Hoàn tất 100%"],
    ["REQ-2026-002", "2026-07-26 14:00", "Marketing", "DM_TAG_BONUS_NIGHT_20K", 450, "Lê QM Spec", "TAGGED_SUCCESS", "2026-07-26 16:45", 448, "=I6/E6", "BATCH_QM_88425 - 2 TX bị khóa tài khoản"],
    ["REQ-2026-004", "2026-07-28 10:00", "Customer Service", "DM_TAG_CS_VIP_5STAR", 30, "Vũ QM Spec", "PROCESSING", "-", 0, "=I7/E7", "Đang push qua Tool Admin Portal"],
    ["REQ-2026-006", "2026-07-29 08:30", "Fleet Ops", "DM_TAG_HUB_SOC_NIGHT", 200, "Phan QM Spec", "PENDING_QM", "-", 0, "=I8/E8", "Mới tiếp nhận từ DM"]
]

for row_idx, row_data in enumerate(sample_data_ws3, 5):
    ws3.row_dimensions[row_idx].height = 22
    for col_idx, val in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.font = REGULAR_FONT
        cell.border = THIN_BORDER
        if col_idx in [1, 2, 6, 7, 8]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx in [5, 9, 10]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if col_idx == 10:
                cell.number_format = "0.0%"
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")


# -------------------------------------------------------------
# TAB 4: 04_TAG_DICTIONARY_METRICS
# -------------------------------------------------------------
ws4 = wb.create_sheet(title="04_TAG_DICTIONARY_METRICS")
ws4.views.sheetView[0].showGridLines = True

ws4.merge_cells("A1:G1")
ws4["A1"] = "DANH MỤC TAG CHUẨN HOÁ & BÁO CÁO HIỆU SUẤT (KPI METRICS)"
ws4["A1"].font = TITLE_FONT
ws4["A1"].alignment = Alignment(vertical="center")

# KPI Summary Cards Area
ws4.merge_cells("A3:B3")
ws4["A3"] = "📊 TỔNG QUAN HIỆU SUẤT (SUMMARY METRICS)"
ws4["A3"].font = BOLD_FONT

kpi_labels = [
    ("Tổng Số Request Tiếp Nhận:", "=COUNTA('01_ALL_TAG_REQUESTS'!A5:A500)"),
    ("Tỷ Lệ DM Duyệt (Approval Rate):", "=COUNTIF('01_ALL_TAG_REQUESTS'!L5:L500, \"APPROVED\") / A4"),
    ("Tỷ Lệ QM Tag Thành Công:", "=COUNTIF('01_ALL_TAG_REQUESTS'!R5:R500, \"TAGGED_SUCCESS\") / A4"),
    ("Thời Gian DM Review Trung Bình (SLA):", "2.1 Giờ (Mục tiêu <= 4h)"),
    ("Thời Gian QM Execution Trung Bình:", "1.8 Giờ (Mục tiêu <= 2h)")
]

for idx, (k_lbl, k_val) in enumerate(kpi_labels, 4):
    ws4.cell(row=idx, column=1, value=k_lbl).font = BOLD_FONT
    cell_val = ws4.cell(row=idx, column=2, value=k_val)
    cell_val.font = Font(name="Arial", size=10, bold=True, color="0F766E")
    if "%" in k_lbl or "Tỷ Lệ" in k_lbl:
        cell_val.number_format = "0.0%"

# Tag Dictionary Header
ws4.cell(row=10, column=1, value="📚 TỪ ĐIỂN MÃ TAG HỆ THỐNG (SYSTEM TAG DICTIONARY)").font = TITLE_FONT

headers_dict = ["Mã Tag (Tag Code)", "Tên Tag Hiển Thị", "Nhóm Tag", "Cấp Thẩm Quyền Duyệt", "Thời Hạn Mặc Định", "Mô Tả Mục Đích", "Số Lượng TX Đang Gán"]
ws4.row_dimensions[11].height = 26
for col_num, title in enumerate(headers_dict, 1):
    cell = ws4.cell(row=11, column=col_num, value=title)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")

sample_tags = [
    ["DM_TAG_HUB_LH_VIP_V1", "Hub Offload Linehaul VIP", "Priority Dispatch", "DM Lead", "30 ngày", "Ưu tiên gán đơn chạy tuyến SOC Linehaul", 150],
    ["DM_TAG_BONUS_NIGHT_20K", "Thưởng Ca Đêm 20k", "Incentive Campaign", "DM Lead + Marketing", "Theo Campaign", "Tài xế tích cực ca 18h-22h", 448],
    ["DM_TAG_CS_VIP_5STAR", "Chăm Sóc Khách VIP", "Special Training", "DM Lead + CS Lead", "90 ngày", "Gán đơn cho khách hàng Doanh nghiệp VIP", 30],
    ["DM_TAG_ZONE_RESTRICT", "Giới Hạn Vùng Tân Bình", "Area Restriction", "DM Lead + Hub Lead", "60 ngày", "TX hoàn tất bài test điều phối vùng", 0],
    ["DM_TAG_RISK_SUSPEND", "Tạm Đình Chỉ Khảo Sát", "Penalty / Block", "DM Lead + Risk Lead", "14 ngày", "TX có nghi vấn vi phạm quy trình giao nhận", 12]
]

for row_idx, row_data in enumerate(sample_tags, 12):
    ws4.row_dimensions[row_idx].height = 22
    for col_idx, val in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=val)
        cell.font = REGULAR_FONT
        cell.border = THIN_BORDER
        if col_idx in [1, 3, 4, 5]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 7:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")


# Auto-fit column widths across all sheets
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip merged wide titles in length calculation
            if cell.row in [1, 2]:
                continue
            if cell.value is not None:
                val_str = str(cell.value)
                if len(val_str) > max_len:
                    max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 15)

# Save Workbook
output_path = "/Users/ts-1148/Desktop/Pulu-workspace/Output/Ahamove/04. OPS_METRICS/2026-07-DM-QM-Tag-Management-System.xlsx"
wb.save(output_path)
print(f"Excel file created successfully at: {output_path}")
