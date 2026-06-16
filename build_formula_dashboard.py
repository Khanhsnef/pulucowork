"""
Weekly Ops Dashboard — Full rebuild
Sections: KPI, 6-Week Trend, Region x Service/Segment, Service x Region/Segment,
          Cancel Analysis (6 sub-tables), Active Driver, Suspended, Hourly Performance, DFD Trend
"""
import pandas as pd, warnings, os, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
warnings.filterwarnings('ignore')

SRC = '/Users/ts-1148/Downloads/[S&P] Supply data report - Weekly.xlsx'
OUT = '/Users/ts-1148/Desktop/Pulu-workspace/output/Ahamove/04. OPS_METRICS/weekly-ops-dashboard.xlsx'

# ── COLORS ────────────────────────────────────────────────────────────────────
BLUE='0E4174'; ORANGE='FF7F32'; GREEN='10B981'; RED='EF4444'
WHITE='FFFFFF'; BL='EEF3FB'; DARK='1E293B'; BDR='CBD5E1'
GRAY='94A3B8'; YELLOW='FFF9C4'; PURPLE='6D28D9'; ROSE='BE123C'
TEAL='0D9488'; LBLUE='DBEAFE'; LGREEN='D1FAE5'; LRED='FEE2E2'
HBDR='94A3B8'; NAVY='0F2744'

# ── SHEET REFERENCES ──────────────────────────────────────────────────────────
SVC="'SVC DATA'";  DFD="'DFD DATA'";  DRV="'DRIVER DATA'"
CAN="'CANCEL DATA'"; HRL="'HOURLY DATA'"; SUS="'SUSPEND DATA'"; DRC="'DRV CANCEL DATA'"

SVC_N=2000; DFD_N=500; DRV_N=1000; CAN_N=5000
HRL_N=15000; SUS_N=500; DRC_N=500

REGIONS=['SGN','HAN','EXP','NW']
SERVICES=['1H','2H','4H','BULKY','AIFOOD']
SEGS=['FT','NIM','NLM','PT','Return']
CITIES=['SGN','HAN','EXP']   # driver/suspend data (no NW)

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
def mf(h): return PatternFill('solid', fgColor=h)
def fn(bold=False,color=DARK,size=9,italic=False):
    return Font(bold=bold,color=color,size=size,name='Arial',italic=italic)
def al(h='center',v='center'): return Alignment(horizontal=h,vertical=v)
def mb(c=BDR): s=Side(border_style='thin',color=c); return Border(left=s,right=s,top=s,bottom=s)
def mb_hdr(): s=Side(border_style='thin',color=BLUE); return Border(left=s,right=s,top=s,bottom=s)

# ── DATA HELPERS ──────────────────────────────────────────────────────────────
def to_pydate(v):
    if pd.isna(v): return None
    return pd.Timestamp(v).date()

def safe_int(v): return int(v) if pd.notna(v) else 0
def safe_float(v): return float(v) if pd.notna(v) else 0.0
def safe_str(v): return str(v) if pd.notna(v) else ''

# ── FORMULA HELPERS ───────────────────────────────────────────────────────────
# Single-criteria SUMPRODUCT (period + 1 filter)
def sp1(sheet, col, p, fc, fv, n):
    return (f"SUMPRODUCT(({sheet}!$B$4:$B${n}={p})"
            f"*({sheet}!${fc}$4:${fc}${n}=\"{fv}\"),"
            f"{sheet}!${col}$4:${col}${n})")

# Two-criteria SUMPRODUCT (period + 2 filters)
def sp2(sheet, col, p, fc, fv, gc, gv, n):
    return (f"SUMPRODUCT(({sheet}!$B$4:$B${n}={p})"
            f"*({sheet}!${fc}$4:${fc}${n}=\"{fv}\")"
            f"*({sheet}!${gc}$4:${gc}${n}=\"{gv}\"),"
            f"{sheet}!${col}$4:${col}${n})")

# Three-criteria SUMPRODUCT
def sp3(sheet, col, p, fc, fv, gc, gv, hc, hv, n):
    return (f"SUMPRODUCT(({sheet}!$B$4:$B${n}={p})"
            f"*({sheet}!${fc}$4:${fc}${n}=\"{fv}\")"
            f"*({sheet}!${gc}$4:${gc}${n}=\"{gv}\")"
            f"*({sheet}!${hc}$4:${hc}${n}=\"{hv}\"),"
            f"{sheet}!${col}$4:${col}${n})")

# SVC helpers (C=Region D=Service E=Req F=Acc G=Cmp)
def svc_ar1(p, reg):   return f"=IFERROR({sp1(SVC,'F',p,'C',reg,SVC_N)}/{sp1(SVC,'E',p,'C',reg,SVC_N)},0)"
def svc_fr1(p, reg):   return f"=IFERROR({sp1(SVC,'G',p,'C',reg,SVC_N)}/{sp1(SVC,'F',p,'C',reg,SVC_N)},0)"
def svc_col1(col,p,reg): return f"={sp1(SVC,col,p,'C',reg,SVC_N)}"
def svc_ar2(p,reg,svc): return f"=IFERROR({sp2(SVC,'F',p,'C',reg,'D',svc,SVC_N)}/{sp2(SVC,'E',p,'C',reg,'D',svc,SVC_N)},0)"
def svc_fr2(p,reg,svc): return f"=IFERROR({sp2(SVC,'G',p,'C',reg,'D',svc,SVC_N)}/{sp2(SVC,'F',p,'C',reg,'D',svc,SVC_N)},0)"
def svc_col2(col,p,reg,svc): return f"={sp2(SVC,col,p,'C',reg,'D',svc,SVC_N)}"
def svc_ar_svc1(p,svc): return f"=IFERROR({sp1(SVC,'F',p,'D',svc,SVC_N)}/{sp1(SVC,'E',p,'D',svc,SVC_N)},0)"
def svc_fr_svc1(p,svc): return f"=IFERROR({sp1(SVC,'G',p,'D',svc,SVC_N)}/{sp1(SVC,'F',p,'D',svc,SVC_N)},0)"
def svc_col_svc1(col,p,svc): return f"={sp1(SVC,col,p,'D',svc,SVC_N)}"

# DRIVER helpers (C=City D=Segment E=TotalDrv F=GDR G=OnlineHrs H=BusyHrs
#                 I=AcceptOrd J=Reward K=OrderIncome L=AA M=TotalNoti N=AccNoti O=TotalSTP)
def drv1(col,p,city): return f"={sp1(DRV,col,p,'C',city,DRV_N)}"
def drv2(col,p,city,seg): return f"={sp2(DRV,col,p,'C',city,'D',seg,DRV_N)}"
def drv_ar1(p,city):     return f"=IFERROR({sp1(DRV,'N',p,'C',city,DRV_N)}/{sp1(DRV,'M',p,'C',city,DRV_N)},0)"
def drv_fr1(p,city):     return f"=IFERROR({sp1(DRV,'O',p,'C',city,DRV_N)}/{sp1(DRV,'I',p,'C',city,DRV_N)},0)"
def drv_ar2(p,city,seg): return f"=IFERROR({sp2(DRV,'N',p,'C',city,'D',seg,DRV_N)}/{sp2(DRV,'M',p,'C',city,'D',seg,DRV_N)},0)"
def drv_fr2(p,city,seg): return f"=IFERROR({sp2(DRV,'O',p,'C',city,'D',seg,DRV_N)}/{sp2(DRV,'I',p,'C',city,'D',seg,DRV_N)},0)"
def drv_seg_ar(p,seg):   return f"=IFERROR({sp1(DRV,'N',p,'D',seg,DRV_N)}/{sp1(DRV,'M',p,'D',seg,DRV_N)},0)"
def drv_seg_fr(p,seg):   return f"=IFERROR({sp1(DRV,'O',p,'D',seg,DRV_N)}/{sp1(DRV,'I',p,'D',seg,DRV_N)},0)"
def drv_seg1(col,p,seg): return f"={sp1(DRV,col,p,'D',seg,DRV_N)}"
def drv_income(p,city=None,seg=None):
    if city and seg:
        o=sp2(DRV,'K',p,'C',city,'D',seg,DRV_N); r=sp2(DRV,'J',p,'C',city,'D',seg,DRV_N)
        d=sp2(DRV,'E',p,'C',city,'D',seg,DRV_N)
    elif city:
        o=sp1(DRV,'K',p,'C',city,DRV_N); r=sp1(DRV,'J',p,'C',city,DRV_N); d=sp1(DRV,'E',p,'C',city,DRV_N)
    elif seg:
        o=sp1(DRV,'K',p,'D',seg,DRV_N); r=sp1(DRV,'J',p,'D',seg,DRV_N); d=sp1(DRV,'E',p,'D',seg,DRV_N)
    else:
        o=f"SUMIF({DRV}!$B:$B,{p},{DRV}!$K:$K)"; r=f"SUMIF({DRV}!$B:$B,{p},{DRV}!$J:$J)"
        d=f"SUMIF({DRV}!$B:$B,{p},{DRV}!$E:$E)"
    return f"=IFERROR(({o}+{r})/{d},0)"

# CANCEL helpers (C=Reason D=Type E=Service F=City G=Segment H=TotalCancel)
def can1(p,fc,fv):   return f"={sp1(CAN,'H',p,fc,fv,CAN_N)}"
def can2(p,fc,fv,gc,gv): return f"={sp2(CAN,'H',p,fc,fv,gc,gv,CAN_N)}"
def can3(p,fc,fv,gc,gv,hc,hv): return f"={sp3(CAN,'H',p,fc,fv,gc,gv,hc,hv,CAN_N)}"

# HOURLY helpers (B=Period C=OrderDate D=Hour E=Region F=Service G=Req H=Acc I=Cmp J=AR K=FR L=SurgeTotal M=SurgeOrd)
def hrl(col, p, hr_ref, reg=None, svc=None):
    b=f"{HRL}!$B$4:$B${HRL_N}"; d=f"{HRL}!$D$4:$D${HRL_N}"
    e=f"{HRL}!$E$4:$E${HRL_N}"; f=f"{HRL}!$F$4:$F${HRL_N}"
    m=f"{HRL}!${col}$4:${col}${HRL_N}"
    flt=f"({b}={p})*({d}={hr_ref})"
    if reg: flt+=f"*({e}=\"{reg}\")"
    if svc: flt+=f"*({f}=\"{svc}\")"
    return f"SUMPRODUCT({flt},{m})"
def hrl_ar(p,hr,reg=None,svc=None): return f"=IFERROR({hrl('H',p,hr,reg,svc)}/{hrl('G',p,hr,reg,svc)},0)"
def hrl_fr(p,hr,reg=None,svc=None): return f"=IFERROR({hrl('I',p,hr,reg,svc)}/{hrl('H',p,hr,reg,svc)},0)"
def hrl_surge(p,hr,reg=None,svc=None): return f"=IFERROR({hrl('M',p,hr,reg,svc)}/{hrl('G',p,hr,reg,svc)},0)"

# ── LOAD RAW DATA ─────────────────────────────────────────────────────────────
raw = pd.ExcelFile(SRC).parse('data import', header=None)

def to_num(df, cols):
    for c in cols: df[c]=pd.to_numeric(df[c],errors='coerce')
    return df

# SVC
svc_df = raw.iloc[2:,0:8].copy()
svc_df.columns=['period','region','service_type','requested','accept','completed','ar','fr']
svc_df = svc_df.dropna(subset=['period','region'])
svc_df['period']=pd.to_datetime(svc_df['period'])
svc_df = to_num(svc_df,['requested','accept','completed','ar','fr'])
svc_df = svc_df[svc_df['region'].isin(REGIONS)].sort_values(['period','region','service_type'],ascending=[False,True,True])

# DFD
dfd_df = raw.iloc[2:,[0,1,9,10,11,12,17]].copy()
dfd_df.columns=['period','region','total_dfd','referral','digital_ads','others','organic']
dfd_df = dfd_df.dropna(subset=['total_dfd'])
dfd_df['period']=pd.to_datetime(dfd_df['period'])
dfd_df = to_num(dfd_df,['total_dfd','referral','digital_ads','others','organic'])
dfd_df = dfd_df.sort_values(['period','region'],ascending=[False,True])

# DRIVER (expanded with busy_hrs, order_income, total_noti, accept_noti, total_stp)
drv_df = raw.iloc[2:,[82,86,68,89,87,73,84,74,78,85,81,80,76,70]].copy()
drv_df.columns=['period','city','segment','total_driver','gdr_driver','online_hrs','busy_hrs',
                'accept_ord','reward_income','order_income','aa_driver','total_noti','accept_noti','total_stp']
drv_df = drv_df.dropna(subset=['period','city','segment'])
drv_df['period']=pd.to_datetime(drv_df['period'])
drv_df = to_num(drv_df,['total_driver','gdr_driver','online_hrs','busy_hrs','accept_ord',
                         'reward_income','order_income','aa_driver','total_noti','accept_noti','total_stp'])
drv_df = drv_df[drv_df['segment'].isin(SEGS)].sort_values(['period','city','segment'],ascending=[False,True,True])

# CANCEL
can_df = raw.iloc[2:,91:98].copy()
can_df.columns=['period','type_','reason_type','service_type','city_id','segment','total_cancel']
can_df = can_df.dropna(subset=['period','type_'])
can_df['period']=pd.to_datetime(can_df['period'])
can_df = to_num(can_df,['total_cancel'])
can_df = can_df.sort_values(['period','total_cancel'],ascending=[False,False])

# HOURLY
hrl_df = raw.iloc[2:,[110,107,102,105,108,98,104,101,109,103,99,106]].copy()
hrl_df.columns=['period','order_date','hour','region','service','requested','accept','completed','ar','fr','surge_total','surge_order']
hrl_df = hrl_df.dropna(subset=['period','hour','region'])
hrl_df = hrl_df[pd.to_numeric(hrl_df['hour'],errors='coerce').notna()]
hrl_df['hour']=pd.to_numeric(hrl_df['hour'])
hrl_df['period']=pd.to_datetime(hrl_df['period'])
hrl_df['order_date']=pd.to_datetime(hrl_df['order_date'])
hrl_df = to_num(hrl_df,['requested','accept','completed','ar','fr','surge_total','surge_order'])
hrl_df = hrl_df[hrl_df['region'].isin(REGIONS)].sort_values(['period','hour','region'],ascending=[False,True,True])

# SUSPEND
sus_df = raw.iloc[2:,[123,124,125,126]].copy()
sus_df.columns=['period','city','reason','cases']
sus_df = sus_df.dropna(subset=['period','city','reason'])
sus_df = sus_df[sus_df['reason']!='Suspend Flow']
sus_df['period']=pd.to_datetime(sus_df['period'])
sus_df = to_num(sus_df,['cases'])
sus_df = sus_df.sort_values(['period','cases'],ascending=[False,False])

# DRV CANCEL
drc_df = raw.iloc[2:,[118,119,117,113,116,114,115,120,121]].copy()
drc_df.columns=['period','city','service','driver_cancel','accept','complete','cr','cancel_poc','cr_rule_poc']
drc_df = drc_df.dropna(subset=['period','city','service'])
drc_df = drc_df[drc_df['service'].isin(SERVICES)]
drc_df['period']=pd.to_datetime(drc_df['period'])
drc_df = to_num(drc_df,['driver_cancel','accept','complete','cr','cancel_poc','cr_rule_poc'])
drc_df = drc_df.sort_values(['period','city','service'],ascending=[False,True,True])

print(f"SVC:{len(svc_df)} DFD:{len(dfd_df)} DRV:{len(drv_df)} CAN:{len(can_df)} HRL:{len(hrl_df)} SUS:{len(sus_df)} DRC:{len(drc_df)}")

# ── WORKBOOK ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# DATA SHEETS
# ─────────────────────────────────────────────────────────────────────────────
def make_data_sheet(sheet, title, headers, col_widths, fmt_map, rows, note):
    sheet.sheet_view.showGridLines = False
    for col,w in zip('ABCDEFGHIJKLMNOP', col_widths): sheet.column_dimensions[col].width = w
    sheet.row_dimensions[1].height = 8
    sheet.row_dimensions[2].height = 30
    n = len(headers)+1
    sheet.merge_cells(f'B2:{chr(65+n)}2')
    for c in range(2,n+2): sheet.cell(row=2,column=c).fill = mf(BLUE)
    sheet['B2'] = title
    sheet['B2'].font = fn(bold=True,color=WHITE,size=11); sheet['B2'].alignment = al('left')
    sheet.row_dimensions[3].height = 22
    for ci,h in enumerate(headers,2):
        c = sheet.cell(row=3,column=ci,value=h)
        c.font=fn(bold=True,color=WHITE,size=9); c.fill=mf(BLUE); c.alignment=al(); c.border=mb(BLUE)
    for i,rv in enumerate(rows):
        row=4+i; sheet.row_dimensions[row].height = 16
        fc = BL if i%2==0 else WHITE
        for ci,(val,fmt) in enumerate(zip(rv,[fmt_map.get(j,'@') for j in range(2,len(rv)+2)]),2):
            c = sheet.cell(row=row,column=ci,value=val)
            c.font=fn(size=9); c.fill=mf(fc)
            c.number_format=fmt; c.alignment=al('left' if ci==3 else 'center'); c.border=mb()
    sheet.auto_filter.ref=f'B3:{chr(65+n)}{3+len(rows)}'
    sheet.freeze_panes='B4'
    nr=4+len(rows)+2
    sheet.merge_cells(f'B{nr}:{chr(65+n)}{nr}')
    sheet.cell(row=nr,column=2,value=note).font=fn(italic=True,color=GRAY,size=8)
    sheet.cell(row=nr,column=2).alignment=al('left')

# SVC DATA
ws_svc = wb.active; ws_svc.title='SVC DATA'; ws_svc.sheet_properties.tabColor=ORANGE
svc_rows=[]
for _,r in svc_df.iterrows():
    svc_rows.append([to_pydate(r['period']),safe_str(r['region']),safe_str(r['service_type']),
                     safe_int(r['requested']),safe_int(r['accept']),safe_int(r['completed']),
                     safe_float(r['ar']),safe_float(r['fr'])])
make_data_sheet(ws_svc,'SVC DATA — Paste new weekly data here (replace from row 4 down)',
    ['Period','Region','Service','Requested','Accept','Completed','AR %','FR %'],
    [2,14,10,12,14,14,14,10,10],
    {2:'DD/MM/YYYY',3:'@',4:'@',5:'#,##0',6:'#,##0',7:'#,##0',8:'0.0%',9:'0.0%'},
    svc_rows,'UPDATE: Xóa từ row 4 → Paste data mới. Period = date value. AR/FR = decimal 0.875.')

# DFD DATA
ws_dfd=wb.create_sheet('DFD DATA'); ws_dfd.sheet_properties.tabColor=GREEN
dfd_rows=[]
for _,r in dfd_df.iterrows():
    dfd_rows.append([to_pydate(r['period']),safe_str(r['region']),
                     safe_int(r['total_dfd']),safe_int(r['organic']),
                     safe_int(r['referral']),safe_int(r['digital_ads'])])
make_data_sheet(ws_dfd,'DFD DATA — Paste new weekly DFD data here (replace from row 4 down)',
    ['Period','Region','Total DFD','Organic','Referral','Digital Ads'],
    [2,14,10,12,12,12,14],
    {2:'DD/MM/YYYY',3:'@',4:'#,##0',5:'#,##0',6:'#,##0',7:'#,##0'},
    dfd_rows,'UPDATE: Xóa từ row 4 → Paste data mới. Period = date value.')

# DRIVER DATA (expanded)
ws_drv=wb.create_sheet('DRIVER DATA'); ws_drv.sheet_properties.tabColor='1D4ED8'
drv_rows=[]
for _,r in drv_df.iterrows():
    drv_rows.append([
        to_pydate(r['period']),safe_str(r['city']),safe_str(r['segment']),
        safe_int(r['total_driver']),safe_int(r['gdr_driver']),
        round(safe_float(r['online_hrs']),1),round(safe_float(r['busy_hrs']),1),
        safe_int(r['accept_ord']),safe_int(r['reward_income']),safe_int(r['order_income']),
        safe_int(r['aa_driver']),safe_int(r['total_noti']),safe_int(r['accept_noti']),safe_int(r['total_stp'])
    ])
make_data_sheet(ws_drv,'DRIVER DATA — Paste new weekly driver data here (replace from row 4 down)',
    ['Period','City','Segment','Total Driver','GDR Driver','Online Hrs','Busy Hrs',
     'Accept Ord','Reward (VND)','Order Income','AA Driver','Total Noti','Accept Noti','Total STP'],
    [2,14,10,10,12,12,12,12,12,14,14,10,10,10],
    {2:'DD/MM/YYYY',3:'@',4:'@',5:'#,##0',6:'#,##0',7:'#,##0.0',8:'#,##0.0',
     9:'#,##0',10:'#,##0',11:'#,##0',12:'#,##0',13:'#,##0',14:'#,##0',15:'#,##0'},
    drv_rows,'DRIVER DATA cols: B=Period C=City D=Segment E=TotalDrv F=GDR G=OnlineHrs H=BusyHrs I=AccOrd J=Reward K=OrderInc L=AA M=TotalNoti N=AccNoti O=TotalSTP')

# CANCEL DATA
ws_can=wb.create_sheet('CANCEL DATA'); ws_can.sheet_properties.tabColor=RED
can_rows=[]
for i,(_,r) in enumerate(can_df.iterrows()):
    if i>=4000: break
    can_rows.append([to_pydate(r['period']),safe_str(r['type_'])[:60],safe_str(r['reason_type']),
                     safe_str(r['service_type']),safe_str(r['city_id']),
                     safe_str(r['segment']),safe_int(r['total_cancel'])])
make_data_sheet(ws_can,'CANCEL DATA — Paste new weekly cancel data here (replace from row 4 down)',
    ['Period','Cancel Reason','Type','Service','City','Segment','Total Cancel'],
    [2,14,40,12,10,10,12,12],
    {2:'DD/MM/YYYY',3:'@',4:'@',5:'@',6:'@',7:'@',8:'#,##0'},
    can_rows,'UPDATE: Xóa từ row 4 → Paste data mới. C=Cancel Reason, D=Type, E=Service, F=City, G=Segment.')

# HOURLY DATA
ws_hrl=wb.create_sheet('HOURLY DATA'); ws_hrl.sheet_properties.tabColor='7C3AED'
hrl_rows=[]
for _,r in hrl_df.iterrows():
    hrl_rows.append([to_pydate(r['period']),to_pydate(r['order_date']),
                     safe_int(r['hour']),safe_str(r['region']),safe_str(r['service']),
                     safe_int(r['requested']),safe_int(r['accept']),safe_int(r['completed']),
                     safe_float(r['ar']),safe_float(r['fr']),
                     safe_float(r['surge_total']),safe_int(r['surge_order'])])
make_data_sheet(ws_hrl,'HOURLY DATA — Paste new weekly hourly data here (replace from row 4 down)',
    ['Period','Order Date','Hour','Region','Service','Requested','Accept','Completed','AR%','FR%','Surge Total','Surge Orders'],
    [2,14,12,6,10,10,12,12,12,10,10,12,12],
    {2:'DD/MM/YYYY',3:'DD/MM/YYYY',4:'#,##0',5:'@',6:'@',7:'#,##0',8:'#,##0',
     9:'#,##0',10:'0.0%',11:'0.0%',12:'#,##0.0',13:'#,##0'},
    hrl_rows,'HOURLY DATA: B=Period(weekly) C=OrderDate D=Hour(0-23) E=Region F=Service G=Req H=Acc I=Cmp J=AR K=FR L=SurgeTotal M=SurgeOrd')

# SUSPEND DATA
ws_sus=wb.create_sheet('SUSPEND DATA'); ws_sus.sheet_properties.tabColor='64748B'
sus_rows=[]
for _,r in sus_df.iterrows():
    sus_rows.append([to_pydate(r['period']),safe_str(r['city']),safe_str(r['reason']),safe_int(r['cases'])])
make_data_sheet(ws_sus,'SUSPEND DATA — Paste new weekly suspend data here (replace from row 4 down)',
    ['Period','City','Reason (Suspend Flow)','Cases'],
    [2,14,10,40,12],
    {2:'DD/MM/YYYY',3:'@',4:'@',5:'#,##0'},
    sus_rows,'SUSPEND DATA: B=Period C=City D=Reason E=Cases')

# DRV CANCEL DATA
ws_drc=wb.create_sheet('DRV CANCEL DATA'); ws_drc.sheet_properties.tabColor='BE123C'
drc_rows=[]
for _,r in drc_df.iterrows():
    drc_rows.append([to_pydate(r['period']),safe_str(r['city']),safe_str(r['service']),
                     safe_int(r['driver_cancel']),safe_int(r['accept']),safe_int(r['complete']),
                     safe_float(r['cr']),safe_int(r['cancel_poc']),safe_float(r['cr_rule_poc'])])
make_data_sheet(ws_drc,'DRV CANCEL DATA — Driver cancel by service (replace from row 4 down)',
    ['Period','City','Service','Driver Cancel','Accept','Complete','CR%','Cancel POC','CR Rule POC'],
    [2,14,10,12,12,12,12,10,12,12],
    {2:'DD/MM/YYYY',3:'@',4:'@',5:'#,##0',6:'#,##0',7:'#,##0',8:'0.0%',9:'#,##0',10:'0.0%'},
    drc_rows,'DRV CANCEL: B=Period C=City D=Service E=DrvCancel F=Accept G=Complete H=CR I=CancelPOC J=CRRulePOC')

# ─────────────────────────────────────────────────────────────────────────────
# WEEKLY DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet('WEEKLY DASHBOARD', 0)
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = BLUE

for col,w in zip('ABCDEFGHIJKLMNOPQR',[2,12,12,12,12,12,12,12,12,12,12,12,12,12,12,12,2,2]):
    ws.column_dimensions[col].width = w

# ── PARAMS (rows 1-5, hidden) ─────────────────────────────────────────────────
# Row 2: SVC week chain B:G, DFD week chain H:M
ws['B2']="=MAX('SVC DATA'!$B:$B)"; ws['B2'].number_format='DD/MM/YYYY'
for i,(lc,cc) in enumerate([('C','$B$2'),('D','$B$2,$C$2'),('E','$B$2,$C$2,$D$2'),
                              ('F','$B$2,$C$2,$D$2,$E$2'),('G','$B$2,$C$2,$D$2,$E$2,$F$2')],1):
    parts=','.join([f"COUNTIF('SVC DATA'!$B:$B,${x}$2)" for x in cc.replace('$','').replace('2','').split(',')])
    ws[f'{lc}2']=f"=IFERROR(LARGE('SVC DATA'!$B:$B,{parts}+1),\"\")"; ws[f'{lc}2'].number_format='DD/MM/YYYY'

ws['H2']="=MAX('DFD DATA'!$B:$B)"; ws['H2'].number_format='DD/MM/YYYY'
for i,(lc,cc) in enumerate([('I','$H$2'),('J','$H$2,$I$2'),('K','$H$2,$I$2,$J$2'),
                              ('L','$H$2,$I$2,$J$2,$K$2'),('M','$H$2,$I$2,$J$2,$K$2,$L$2')],1):
    parts=','.join([f"COUNTIF('DFD DATA'!$B:$B,${x}$2)" for x in cc.replace('$','').replace('2','').split(',')])
    ws[f'{lc}2']=f"=IFERROR(LARGE('DFD DATA'!$B:$B,{parts}+1),\"\")"; ws[f'{lc}2'].number_format='DD/MM/YYYY'

# Row 3: KPI delta pre-computes
ws['B3']=f"=SUMIF({SVC}!$B:$B,$B$2,{SVC}!$E:$E)"   # W1 Req
ws['C3']=f"=SUMIF({SVC}!$B:$B,$C$2,{SVC}!$E:$E)"   # W2 Req
ws['D3']=f"=IFERROR(SUMIF({SVC}!$B:$B,$B$2,{SVC}!$F:$F)/$B3,0)"  # W1 AR
ws['E3']=f"=IFERROR(SUMIF({SVC}!$B:$B,$C$2,{SVC}!$F:$F)/$C3,0)"  # W2 AR
ws['F3']=f"=IFERROR(SUMIF({SVC}!$B:$B,$B$2,{SVC}!$G:$G)/SUMIF({SVC}!$B:$B,$B$2,{SVC}!$F:$F),0)" # W1 FR
ws['G3']=f"=IFERROR(SUMIF({SVC}!$B:$B,$C$2,{SVC}!$G:$G)/SUMIF({SVC}!$B:$B,$C$2,{SVC}!$F:$F),0)" # W2 FR
ws['H3']=f"=SUMIF({SVC}!$B:$B,$B$2,{SVC}!$G:$G)"   # W1 Completed
ws['I3']=f"=SUMIF({SVC}!$B:$B,$C$2,{SVC}!$G:$G)"   # W2 Completed
ws['J3']=f"=SUMIF({DFD}!$B:$B,$H$2,{DFD}!$D:$D)"   # W1 DFD
ws['K3']=f"=SUMIF({DFD}!$B:$B,$I$2,{DFD}!$D:$D)"   # W2 DFD

# Row 4: Latest periods per sheet + total cancel
ws['B4']=f"=MAX({DFD}!$B:$B)"
ws['C4']=f"=MAX({DRV}!$B:$B)"
ws['D4']=f"=MAX({CAN}!$B:$B)"
ws['E4']=f"=SUMIF({CAN}!$B:$B,$D$4,{CAN}!$H:$H)"
ws['F4']=f"=MAX({HRL}!$B:$B)"
ws['G4']=f"=MAX({SUS}!$B:$B)"
ws['H4']=f"=MAX({DRC}!$B:$B)"

# Row 5: Driver total KPI pre-computes (for top cards)
ws['B5']=f"=SUMIF({DRV}!$B:$B,$C$4,{DRV}!$E:$E)"   # total drivers
ws['C5']=f"=SUMIF({DRV}!$B:$B,$C$4,{DRV}!$L:$L)"   # total AA
ws['D5']=f"=SUMIF({DRV}!$B:$B,$C$4,{DRV}!$G:$G)"   # total online hrs
ws['E5']=f"=SUMIF({DRV}!$B:$B,$C$4,{DRV}!$H:$H)"   # total busy hrs
ws['F5']=f"=IFERROR((SUMIF({DRV}!$B:$B,$C$4,{DRV}!$K:$K)+SUMIF({DRV}!$B:$B,$C$4,{DRV}!$J:$J))/$B5,0)" # avg income
ws['G5']=f"=SUMIF({SUS}!$B:$B,$G$4,{SUS}!$E:$E)"   # total suspended

for r in [1,2,3,4,5]: ws.row_dimensions[r].height = 0.1

# ─── DASHBOARD LAYOUT HELPERS ─────────────────────────────────────────────────
def rh(row, h=18): ws.row_dimensions[row].height = h

def section(r, title, color=NAVY):
    rh(r, 20)
    ws.merge_cells(f'B{r}:Q{r}')
    c = ws.cell(row=r, column=2, value=f'  {title}')
    c.font=fn(bold=True,color=WHITE,size=9)
    c.fill=mf(color); c.alignment=al('left')

def subsection(r, title):
    rh(r, 16)
    ws.merge_cells(f'B{r}:Q{r}')
    c = ws.cell(row=r, column=2, value=f'  {title}')
    c.font=fn(bold=True,color=DARK,size=8,italic=True)
    c.fill=mf('E2E8F0'); c.alignment=al('left')

def trow(r, h=17): rh(r, h)

def hcell(r, col, text, span=1):
    if span > 1:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+span-1)
    c = ws.cell(row=r, column=col, value=text)
    c.font=fn(bold=True,color=WHITE,size=8); c.fill=mf(BLUE)
    c.alignment=al(); c.border=mb(BLUE)
    return c

def dcell(r, col, val, fmt='@', bold=False, color=DARK, align='center', bg=None):
    c = ws.cell(row=r, column=col, value=val)
    c.font=fn(bold=bold,color=color,size=9)
    c.fill=mf(bg or BL if (r%2==0) else WHITE)
    c.number_format=fmt; c.alignment=al(align); c.border=mb()
    return c

def row_fill(r, col_start, col_end, bg):
    for co in range(col_start, col_end+1):
        ws.cell(row=r, column=co).fill = mf(bg)

# ─── KPI SNAPSHOT ──────────────────────────────────────────────────────────────
r = 6
rh(r, 8)

r = 7  # Banner
rh(r, 36)
ws.merge_cells(f'B{r}:J{r}')
ws[f'B{r}'] = 'AHAMOVE — WEEKLY OPS DASHBOARD'
ws[f'B{r}'].font = fn(bold=True,color=WHITE,size=18)
ws[f'B{r}'].alignment = al('left')
for co in range(2,18): ws.cell(row=r,column=co).fill = mf(BLUE)
ws.merge_cells(f'K{r}:Q{r}')
ws[f'K{r}'] = f'=IFERROR("W/E "&TEXT($B$2,"DD/MM/YYYY"),"No data")'
ws[f'K{r}'].font = fn(bold=True,color=ORANGE,size=12)
ws[f'K{r}'].alignment = al('right')
for co in range(11,18): ws.cell(row=r,column=co).fill = mf(BLUE)

r = 8  # Subtitle
rh(r, 18)
ws.merge_cells(f'B{r}:Q{r}')
ws[f'B{r}'] = 'Driver Management  |  Bike Instant  |  SGN · HAN · EXP · NW'
ws[f'B{r}'].font = fn(color='AACCEE',size=9,italic=True)
ws[f'B{r}'].alignment = al('left')
for co in range(2,18): ws.cell(row=r,column=co).fill = mf(BLUE)

r = 9; rh(r, 6)

# KPI section
r = 10
section(r, 'KPI SNAPSHOT  —  Latest Week  (auto-refresh from SVC DATA)')

def kpi_card(val_row, sub_row, col_start, title, val_fml, sub_fml, val_fmt, val_color=GREEN):
    rh(val_row-1 if val_row==11 else val_row, 14)
    # title row
    ws.merge_cells(start_row=val_row-1,start_column=col_start,end_row=val_row-1,end_column=col_start+2)
    tc=ws.cell(row=val_row-1,column=col_start,value=title)
    tc.font=fn(bold=True,color=GRAY,size=7); tc.fill=mf(WHITE)
    tc.alignment=al(); tc.border=mb('E2E8F0')
    # value row
    rh(val_row, 26)
    ws.merge_cells(start_row=val_row,start_column=col_start,end_row=val_row,end_column=col_start+2)
    vc=ws.cell(row=val_row,column=col_start,value=val_fml)
    vc.font=fn(bold=True,color=val_color,size=18); vc.fill=mf(WHITE)
    vc.number_format=val_fmt; vc.alignment=al(); vc.border=mb('E2E8F0')
    # sub row
    rh(sub_row, 16)
    ws.merge_cells(start_row=sub_row,start_column=col_start,end_row=sub_row,end_column=col_start+2)
    sc=ws.cell(row=sub_row,column=col_start,value=sub_fml)
    sc.font=fn(color=GRAY,size=8); sc.fill=mf(WHITE)
    sc.alignment=al(); sc.border=mb('E2E8F0')

# Title row for KPI cards
r = 11
rh(r, 14)
for co in range(2, 17):
    ws.cell(row=r, column=co).fill = mf(WHITE)
    ws.cell(row=r, column=co).border = mb('E2E8F0')

kpi_card(12, 13, 2,  'TOTAL REQUESTED',
    f"=SUMIF({SVC}!$B:$B,$B$2,{SVC}!$E:$E)",
    '="vs prev: "&IF($B$3>$C$3,"▲","▼")&TEXT(ABS($B$3-$C$3),"#,##0")',
    '#,##0', GREEN)
kpi_card(12, 13, 5,  'ACCEPTANCE RATE',
    f"=IFERROR(SUMIF({SVC}!$B:$B,$B$2,{SVC}!$F:$F)/$B$3,0)",
    '="vs prev: "&IF($D$3>=$E$3,"▲","▼")&TEXT(ABS($D$3-$E$3)*100,"0.0")&"pp"',
    '0.0%', GREEN)
kpi_card(12, 13, 8,  'FULFILLMENT RATE',
    f"=IFERROR(SUMIF({SVC}!$B:$B,$B$2,{SVC}!$G:$G)/SUMIF({SVC}!$B:$B,$B$2,{SVC}!$F:$F),0)",
    '="vs prev: "&IF($F$3>=$G$3,"▲","▼")&TEXT(ABS($F$3-$G$3)*100,"0.0")&"pp"',
    '0.0%', GREEN)
kpi_card(12, 13, 11, 'COMPLETED ORDERS',
    f"=SUMIF({SVC}!$B:$B,$B$2,{SVC}!$G:$G)",
    '="vs prev: "&IF($H$3>$I$3,"▲","▼")&TEXT(ABS($H$3-$I$3),"#,##0")',
    '#,##0', GREEN)

r = 14
rh(r, 14)
for co in range(2, 17):
    ws.cell(row=r, column=co).fill = mf(WHITE)
    ws.cell(row=r, column=co).border = mb('E2E8F0')

kpi_card(15, 16, 2,  'ACTIVE DRIVERS (AA)',
    '=$C$5',
    '=IFERROR("Period: "&TEXT($C$4,"DD/MM/YYYY"),"")',
    '#,##0', BLUE)
kpi_card(15, 16, 5,  'DFD — NEW DRIVERS',
    f"=SUMIF({DFD}!$B:$B,$H$2,{DFD}!$D:$D)",
    '="vs prev: "&IF($J$3>$K$3,"▲","▼")&TEXT(ABS($J$3-$K$3),"#,##0")',
    '#,##0', ORANGE)
kpi_card(15, 16, 8,  'SGN ACCEPTANCE RATE',
    svc_ar1('$B$2','SGN'),
    '"*Target ≥ 85%"','0.0%', GREEN)
kpi_card(15, 16, 11, 'HAN ACCEPTANCE RATE',
    svc_ar1('$B$2','HAN'),
    '"*Target ≥ 85%"','0.0%', GREEN)

r = 17; rh(r, 6)

# ─── 6-WEEK TREND ─────────────────────────────────────────────────────────────
r = 18; section(r, '6-WEEK PERFORMANCE TREND')
r = 19; trow(r, 20)
for col,txt in zip([2,3,4,5,6,7,8,9],['Week','Requested','Accepted','Completed','AR %','FR %','WoW Req %','WoW AR pp']):
    hcell(r, col, txt)

week_cols = ['$B$2','$C$2','$D$2','$E$2','$F$2','$G$2']
for i,wc in enumerate(week_cols):
    r = 20+i; trow(r, 17)
    bg = YELLOW if i==0 else (BL if i%2==0 else WHITE)
    nwc = week_cols[i+1] if i<5 else None
    req_w = f"=SUMIF({SVC}!$B:$B,{wc},{SVC}!$E:$E)"
    acc_w = f"=SUMIF({SVC}!$B:$B,{wc},{SVC}!$F:$F)"
    cmp_w = f"=SUMIF({SVC}!$B:$B,{wc},{SVC}!$G:$G)"
    ar_w  = f"=IFERROR(SUMIF({SVC}!$B:$B,{wc},{SVC}!$F:$F)/SUMIF({SVC}!$B:$B,{wc},{SVC}!$E:$E),0)"
    fr_w  = f"=IFERROR(SUMIF({SVC}!$B:$B,{wc},{SVC}!$G:$G)/SUMIF({SVC}!$B:$B,{wc},{SVC}!$F:$F),0)"
    if nwc:
        wow_req=f"=IFERROR(SUMIF({SVC}!$B:$B,{wc},{SVC}!$E:$E)/SUMIF({SVC}!$B:$B,{nwc},{SVC}!$E:$E)-1,\"—\")"
        wow_ar =f"=IFERROR((SUMIF({SVC}!$B:$B,{wc},{SVC}!$F:$F)/SUMIF({SVC}!$B:$B,{wc},{SVC}!$E:$E))-(SUMIF({SVC}!$B:$B,{nwc},{SVC}!$F:$F)/SUMIF({SVC}!$B:$B,{nwc},{SVC}!$E:$E)),\"—\")"
    else:
        wow_req='"—"'; wow_ar='"—"'
    vals=[f'=IFERROR(IF({wc}="","","W/E "&TEXT({wc},"DD/MM/YY")),"")',req_w,acc_w,cmp_w,ar_w,fr_w,wow_req,wow_ar]
    fmts=['@','#,##0','#,##0','#,##0','0.0%','0.0%','0.0%','0.00%']
    for val,fmt,col in zip(vals,fmts,[2,3,4,5,6,7,8,9]):
        c=ws.cell(row=r,column=col,value=val)
        c.font=fn(bold=(i==0),color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()

r = 26; rh(r, 6)

# ─── REGION BREAKDOWN ─────────────────────────────────────────────────────────
r = 27; section(r, 'REGION BREAKDOWN  —  Latest Week vs Prev Week')

# Region Summary
r = 28; subsection(r, 'Summary by Region')
r = 29; trow(r, 19)
for col,txt in zip([2,3,4,5,6,7,8,9,10],
                   ['Region','Requested','Completed','AR% Cur','AR% Prev','AR Δ pp','FR% Cur','FR% Prev','FR Δ pp']):
    hcell(r, col, txt)
for i,reg in enumerate(REGIONS):
    r=30+i; trow(r); bg=BL if i%2==0 else WHITE
    ar_c=svc_ar1('$B$2',reg); ar_p=svc_ar1('$C$2',reg)
    fr_c=svc_fr1('$B$2',reg); fr_p=svc_fr1('$C$2',reg)
    ar_d=f"={ar_c[1:]}-{ar_p[1:]}"; fr_d=f"={fr_c[1:]}-{fr_p[1:]}"
    vals=[reg,svc_col1('E','$B$2',reg),svc_col1('G','$B$2',reg),ar_c,ar_p,ar_d,fr_c,fr_p,fr_d]
    fmts=['@','#,##0','#,##0','0.0%','0.0%','+0.00%;-0.00%;"-"','0.0%','0.0%','+0.00%;-0.00%;"-"']
    for val,fmt,col in zip(vals,fmts,[2,3,4,5,6,7,8,9,10]):
        c=ws.cell(row=r,column=col,value=val)
        c.font=fn(bold=(col==2),color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()
r=34; rh(r,4)

# Region x Service
r=35; subsection(r,'Region × Service Breakdown')
r=36; trow(r,19)
for col,txt in zip([2,3,4,5,6,7,8],['Region','Service','Requested','AR%','FR%','Completed','Vol% of Region']):
    hcell(r,col,txt)
cur_r=37
for reg in REGIONS:
    # Region group header
    trow(cur_r, 14)
    ws.merge_cells(start_row=cur_r,start_column=2,end_row=cur_r,end_column=8)
    gc=ws.cell(row=cur_r,column=2,value=f'  {reg}')
    gc.font=fn(bold=True,color=WHITE,size=8); gc.fill=mf(BLUE+'AA' if reg!='SGN' else '1A5FA0')
    gc.alignment=al('left'); gc.border=mb(BLUE)
    for co in range(3,9): ws.cell(row=cur_r,column=co).fill=mf(BLUE+'88')
    cur_r+=1
    reg_req=f"={sp1(SVC,'E','$B$2','C',reg,SVC_N)}"
    for j,svc_s in enumerate(SERVICES):
        trow(cur_r); bg=BL if j%2==0 else WHITE
        req=svc_col2('E','$B$2',reg,svc_s)
        ar=svc_ar2('$B$2',reg,svc_s); fr=svc_fr2('$B$2',reg,svc_s)
        cmp=svc_col2('G','$B$2',reg,svc_s)
        vol=f"=IFERROR({req[1:]}/{reg_req[1:]},0)"
        for val,fmt,col in zip([reg,svc_s,req,ar,fr,cmp,vol],
                                ['@','@','#,##0','0.0%','0.0%','#,##0','0.0%'],
                                [2,3,4,5,6,7,8]):
            c=ws.cell(row=cur_r,column=col,value=val)
            c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
            c.number_format=fmt; c.alignment=al('left' if col<=3 else 'center'); c.border=mb()
        cur_r+=1
rh(cur_r,4); cur_r+=1

# Region x Segment (Driver metrics)
r=cur_r; subsection(r,'Region × Segment — Driver Performance (from DRIVER DATA)')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5,6,7,8,9,10],
                   ['Region/City','Segment','Total Drv','AA','AR drv%','FR drv%','Online Hrs','Busy Hrs','Avg Income']):
    hcell(cur_r,col,txt)
cur_r+=1
for city in CITIES:
    trow(cur_r,14)
    ws.merge_cells(start_row=cur_r,start_column=2,end_row=cur_r,end_column=10)
    gc=ws.cell(row=cur_r,column=2,value=f'  {city}')
    gc.font=fn(bold=True,color=WHITE,size=8); gc.fill=mf(BLUE)
    gc.alignment=al('left'); gc.border=mb(BLUE)
    cur_r+=1
    for j,seg in enumerate(SEGS):
        trow(cur_r); bg=BL if j%2==0 else WHITE
        for val,fmt,col in zip([
            city, seg,
            drv2('E','$C$4',city,seg), drv2('L','$C$4',city,seg),
            drv_ar2('$C$4',city,seg), drv_fr2('$C$4',city,seg),
            drv2('G','$C$4',city,seg), drv2('H','$C$4',city,seg),
            drv_income('$C$4',city,seg)],
            ['@','@','#,##0','#,##0','0.0%','0.0%','#,##0.0','#,##0.0','#,##0'],
            [2,3,4,5,6,7,8,9,10]):
            c=ws.cell(row=cur_r,column=col,value=val)
            c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
            c.number_format=fmt; c.alignment=al('left' if col<=3 else 'center'); c.border=mb()
        cur_r+=1
rh(cur_r,6); cur_r+=1

# ─── SERVICE BREAKDOWN ────────────────────────────────────────────────────────
r=cur_r; section(r,'SERVICE BREAKDOWN  —  Latest Week')
cur_r+=1

# Service Summary
r=cur_r; subsection(r,'Summary by Service')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5,6,7,8],['Service','Requested','Accepted','Completed','AR %','FR %','Vol Mix %']):
    hcell(cur_r,col,txt)
cur_r+=1
tot_req=f"SUMIF({SVC}!$B:$B,$B$2,{SVC}!$E:$E)"
for i,svc_s in enumerate(SERVICES):
    trow(cur_r); bg=BL if i%2==0 else WHITE
    req=svc_col_svc1('E','$B$2',svc_s)
    vol=f"=IFERROR({req[1:]}/({tot_req}),0)"
    for val,fmt,col in zip([svc_s,req,svc_col_svc1('F','$B$2',svc_s),svc_col_svc1('G','$B$2',svc_s),
                            svc_ar_svc1('$B$2',svc_s),svc_fr_svc1('$B$2',svc_s),vol],
                            ['@','#,##0','#,##0','#,##0','0.0%','0.0%','0.0%'],[2,3,4,5,6,7,8]):
        c=ws.cell(row=cur_r,column=col,value=val)
        c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()
    cur_r+=1
rh(cur_r,4); cur_r+=1

# Service x Region
r=cur_r; subsection(r,'Service × Region Breakdown')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5,6,7,8],['Service','Region','Requested','Completed','AR%','FR%','Vol% of Svc']):
    hcell(cur_r,col,txt)
cur_r+=1
for svc_s in SERVICES:
    trow(cur_r,14)
    ws.merge_cells(start_row=cur_r,start_column=2,end_row=cur_r,end_column=8)
    gc=ws.cell(row=cur_r,column=2,value=f'  {svc_s}')
    gc.font=fn(bold=True,color=WHITE,size=8); gc.fill=mf(BLUE)
    gc.alignment=al('left'); gc.border=mb(BLUE)
    cur_r+=1
    svc_req=f"={sp1(SVC,'E','$B$2','D',svc_s,SVC_N)}"
    for j,reg in enumerate(REGIONS):
        trow(cur_r); bg=BL if j%2==0 else WHITE
        req=svc_col2('E','$B$2',reg,svc_s)
        cmp=svc_col2('G','$B$2',reg,svc_s)
        ar=svc_ar2('$B$2',reg,svc_s); fr=svc_fr2('$B$2',reg,svc_s)
        vol=f"=IFERROR({req[1:]}/{svc_req[1:]},0)"
        for val,fmt,col in zip([svc_s,reg,req,cmp,ar,fr,vol],
                                ['@','@','#,##0','#,##0','0.0%','0.0%','0.0%'],[2,3,4,5,6,7,8]):
            c=ws.cell(row=cur_r,column=col,value=val)
            c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
            c.number_format=fmt; c.alignment=al('left' if col<=3 else 'center'); c.border=mb()
        cur_r+=1
rh(cur_r,4); cur_r+=1

# Service x Segment (Driver AR/FR)
r=cur_r; subsection(r,'Service × Segment — note: driver segment metrics aggregated by city; no service split in DRIVER DATA')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5,6,7,8,9],['Segment','Total Drv','AA','AR drv%','FR drv%','Online Hrs','Busy Hrs','Avg Income']):
    hcell(cur_r,col,txt)
cur_r+=1
# Show overall segment summary (aggregated all cities, latest driver period)
for i,seg in enumerate(SEGS):
    trow(cur_r); bg=BL if i%2==0 else WHITE
    for val,fmt,col in zip([
        seg,
        drv_seg1('E','$C$4',seg), drv_seg1('L','$C$4',seg),
        drv_seg_ar('$C$4',seg), drv_seg_fr('$C$4',seg),
        drv_seg1('G','$C$4',seg), drv_seg1('H','$C$4',seg),
        drv_income('$C$4',seg=seg)],
        ['@','#,##0','#,##0','0.0%','0.0%','#,##0.0','#,##0.0','#,##0'],
        [2,3,4,5,6,7,8,9]):
        c=ws.cell(row=cur_r,column=col,value=val)
        c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()
    cur_r+=1
rh(cur_r,6); cur_r+=1

# ─── CANCEL ANALYSIS ──────────────────────────────────────────────────────────
r=cur_r; section(r,'CANCEL ANALYSIS  —  Latest Week (CANCEL DATA + DRV CANCEL DATA)')
cur_r+=1

# Top reasons
r=cur_r; subsection(r,'Top Cancel Reasons  (type_ column)')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5,6,7],['Cancel Reason','SGN','HAN','EXP','NW','Total']):
    hcell(cur_r,col,txt)
cur_r+=1
top_reasons = (can_df[can_df['period']==can_df['period'].max()]
               .groupby('type_')['total_cancel'].sum()
               .sort_values(ascending=False).head(12).index.tolist())
for i,rsn in enumerate(top_reasons):
    trow(cur_r); bg=BL if i%2==0 else WHITE
    rsn_safe = rsn[:50] if rsn else ''
    totf = '+'.join([can3('$D$4','C',rsn_safe,'F',city,'C',rsn_safe,CAN_N)[1:] if False
                     else sp2(CAN,'H','$D$4','C',rsn_safe,'F',city,CAN_N) for city in CITIES+['NW']])
    # Simpler: just per city
    sgn=f"={sp2(CAN,'H','$D$4','C',rsn_safe,'F','SGN',CAN_N)}"
    han=f"={sp2(CAN,'H','$D$4','C',rsn_safe,'F','HAN',CAN_N)}"
    exp=f"={sp2(CAN,'H','$D$4','C',rsn_safe,'F','EXP',CAN_N)}"
    nw=f"={sp2(CAN,'H','$D$4','C',rsn_safe,'F','NW',CAN_N)}"
    total=f"={sp1(CAN,'H','$D$4','C',rsn_safe,CAN_N)}"
    for val,fmt,col in zip([rsn_safe,sgn,han,exp,nw,total],
                            ['@','#,##0','#,##0','#,##0','#,##0','#,##0'],[2,3,4,5,6,7]):
        c=ws.cell(row=cur_r,column=col,value=val)
        c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()
    cur_r+=1
rh(cur_r,4); cur_r+=1

# Cancel by Region
r=cur_r; subsection(r,'Cancel by Region')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5],['Region','Total Cancel','% of All','vs Prev Week']):
    hcell(cur_r,col,txt)
cur_r+=1
total_can_cur=f"SUMIF({CAN}!$B:$B,$D$4,{CAN}!$H:$H)"
for i,reg in enumerate(REGIONS+CITIES[:0]):
    trow(cur_r); bg=BL if i%2==0 else WHITE
    tot=f"={sp1(CAN,'H','$D$4','F',reg,CAN_N)}"
    pct=f"=IFERROR({tot[1:]}/({total_can_cur}),0)"
    prv=f"=IFERROR({tot[1:]}/{sp1(CAN,'H','$E$4','F',reg,CAN_N) if False else sp1(CAN,'H',f'=IFERROR(LARGE({CAN}!$B:$B,COUNTIF({CAN}!$B:$B,$D$4)+1),0)','F',reg,CAN_N)}-1,\"—\")"
    # Simplified prev week cancel
    d4='"$D$4"'
    prev_can=f"SUMPRODUCT(({CAN}!$B$4:$B${CAN_N}=IFERROR(LARGE({CAN}!$B$4:$B${CAN_N},COUNTIF({CAN}!$B$4:$B${CAN_N},$D$4)+1),0))*({CAN}!$F$4:$F${CAN_N}=\"{reg}\"),{CAN}!$H$4:$H${CAN_N})"
    vsw=f"=IFERROR(({tot[1:]})/({prev_can})-1,\"—\")"
    for val,fmt,col in zip([reg,tot,pct,vsw],['@','#,##0','0.0%','0.0%'],[2,3,4,5]):
        c=ws.cell(row=cur_r,column=col,value=val)
        c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()
    cur_r+=1
rh(cur_r,4); cur_r+=1

# Cancel by Service
r=cur_r; subsection(r,'Cancel by Service × Region (pivot)')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5,6,7],['Service','SGN','HAN','EXP','NW','Total']):
    hcell(cur_r,col,txt)
cur_r+=1
for i,svc_s in enumerate(SERVICES):
    trow(cur_r); bg=BL if i%2==0 else WHITE
    sgn=f"={sp2(CAN,'H','$D$4','E',svc_s,'F','SGN',CAN_N)}"
    han=f"={sp2(CAN,'H','$D$4','E',svc_s,'F','HAN',CAN_N)}"
    exp=f"={sp2(CAN,'H','$D$4','E',svc_s,'F','EXP',CAN_N)}"
    nw=f"={sp2(CAN,'H','$D$4','E',svc_s,'F','NW',CAN_N)}"
    tot=f"={sp1(CAN,'H','$D$4','E',svc_s,CAN_N)}"
    for val,fmt,col in zip([svc_s,sgn,han,exp,nw,tot],
                            ['@','#,##0','#,##0','#,##0','#,##0','#,##0'],[2,3,4,5,6,7]):
        c=ws.cell(row=cur_r,column=col,value=val)
        c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()
    cur_r+=1
rh(cur_r,4); cur_r+=1

# Cancel by Segment
r=cur_r; subsection(r,'Cancel by Segment × Service (pivot)')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5,6,7,8],['Segment','1H','2H','4H','BULKY','AIFOOD','Total']):
    hcell(cur_r,col,txt)
cur_r+=1
for i,seg in enumerate(SEGS):
    trow(cur_r); bg=BL if i%2==0 else WHITE
    vals=[seg]
    fmts=['@']
    for svc_s in SERVICES:
        vals.append(f"={sp2(CAN,'H','$D$4','G',seg,'E',svc_s,CAN_N)}")
        fmts.append('#,##0')
    vals.append(f"={sp1(CAN,'H','$D$4','G',seg,CAN_N)}")
    fmts.append('#,##0')
    for val,fmt,col in zip(vals,fmts,[2,3,4,5,6,7,8]):
        c=ws.cell(row=cur_r,column=col,value=val)
        c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()
    cur_r+=1
rh(cur_r,4); cur_r+=1

# Driver Cancel by Service x City
r=cur_r; subsection(r,'Driver Cancel Rate by Service × City  (DRV CANCEL DATA — drivers cancelling after accepting)')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5,6,7,8,9],['Service','City','Driver Cancel','Accept','Complete','CR%','Cancel POC','CR Rule POC']):
    hcell(cur_r,col,txt)
cur_r+=1
for svc_s in SERVICES:
    trow(cur_r,14)
    ws.merge_cells(start_row=cur_r,start_column=2,end_row=cur_r,end_column=9)
    gc=ws.cell(row=cur_r,column=2,value=f'  {svc_s}')
    gc.font=fn(bold=True,color=WHITE,size=8); gc.fill=mf(BLUE)
    gc.alignment=al('left'); gc.border=mb(BLUE)
    cur_r+=1
    for j,city in enumerate(CITIES):
        trow(cur_r); bg=BL if j%2==0 else WHITE
        dcan=f"={sp2(DRC,'E','$H$4','C',city,'D',svc_s,DRC_N)}"
        acc=f"={sp2(DRC,'F','$H$4','C',city,'D',svc_s,DRC_N)}"
        cmp=f"={sp2(DRC,'G','$H$4','C',city,'D',svc_s,DRC_N)}"
        cr=f"=IFERROR({dcan[1:]}/{acc[1:]},0)"
        poc=f"={sp2(DRC,'I','$H$4','C',city,'D',svc_s,DRC_N)}"
        cr_poc=f"=IFERROR({poc[1:]}/{acc[1:]},0)"
        for val,fmt,col in zip([svc_s,city,dcan,acc,cmp,cr,poc,cr_poc],
                                ['@','@','#,##0','#,##0','#,##0','0.0%','#,##0','0.0%'],[2,3,4,5,6,7,8,9]):
            c=ws.cell(row=cur_r,column=col,value=val)
            c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
            c.number_format=fmt; c.alignment=al('left' if col<=3 else 'center'); c.border=mb()
        cur_r+=1
rh(cur_r,6); cur_r+=1

# ─── ACTIVE DRIVER ────────────────────────────────────────────────────────────
r=cur_r; section(r,'ACTIVE DRIVER METRICS  —  Latest DRIVER DATA period  ($C$4)')
cur_r+=1

# Driver KPI Cards
r=cur_r; subsection(r,'KPI Overview  (All Cities, All Segments)')
cur_r+=1
rh(cur_r,14)
for co in range(2,17): ws.cell(row=cur_r,column=co).fill=mf(WHITE); ws.cell(row=cur_r,column=co).border=mb('E2E8F0')

def drv_kpi_mini(val_row, sub_row, col_start, title, val_fml, val_fmt, val_color=BLUE):
    rh(val_row, 24)
    ws.merge_cells(start_row=val_row-1,start_column=col_start,end_row=val_row-1,end_column=col_start+2)
    tc=ws.cell(row=val_row-1,column=col_start,value=title)
    tc.font=fn(bold=True,color=GRAY,size=7); tc.fill=mf(WHITE)
    tc.alignment=al(); tc.border=mb('E2E8F0')
    ws.merge_cells(start_row=val_row,start_column=col_start,end_row=val_row,end_column=col_start+2)
    vc=ws.cell(row=val_row,column=col_start,value=val_fml)
    vc.font=fn(bold=True,color=val_color,size=16); vc.fill=mf(WHITE)
    vc.number_format=val_fmt; vc.alignment=al(); vc.border=mb('E2E8F0')
    rh(sub_row, 14)
    ws.merge_cells(start_row=sub_row,start_column=col_start,end_row=sub_row,end_column=col_start+2)
    sc=ws.cell(row=sub_row,column=col_start,value=f'=IFERROR("Period: "&TEXT($C$4,"DD/MM/YYYY"),"")')
    sc.font=fn(color=GRAY,size=7); sc.fill=mf(WHITE)
    sc.alignment=al(); sc.border=mb('E2E8F0')

drv_kpi_mini(cur_r+1,cur_r+2,2,'TOTAL ACTIVE DRIVERS','=$B$5','#,##0',BLUE)
drv_kpi_mini(cur_r+1,cur_r+2,5,'AA DRIVERS','=$C$5','#,##0',GREEN)
drv_kpi_mini(cur_r+1,cur_r+2,8,'AVG INCOME / DRIVER','=$F$5','#,##0',ORANGE)
drv_kpi_mini(cur_r+1,cur_r+2,11,'ONLINE HOURS (TOTAL)','=$D$5','#,##0.0',BLUE)
cur_r+=3

rh(cur_r,14)
for co in range(2,17): ws.cell(row=cur_r,column=co).fill=mf(WHITE); ws.cell(row=cur_r,column=co).border=mb('E2E8F0')
drv_kpi_mini(cur_r+1,cur_r+2,2,'BUSY HOURS (TOTAL)','=$E$5','#,##0.0',PURPLE)
drv_kpi_mini(cur_r+1,cur_r+2,5,'UTILIZATION (BUSY/ONLINE)',
    '=IFERROR($E$5/$D$5,0)','0.0%',TEAL)
drv_kpi_mini(cur_r+1,cur_r+2,8,'GDR RATE',
    f'=IFERROR(SUMIF({DRV}!$B:$B,$C$4,{DRV}!$F:$F)/$B$5,0)','0.0%',GREEN)
drv_kpi_mini(cur_r+1,cur_r+2,11,'TOTAL SUSPENDED','=$G$5','#,##0',RED)
cur_r+=3
rh(cur_r,4); cur_r+=1

# Driver by City
r=cur_r; subsection(r,'Active Driver by City/Region')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5,6,7,8,9,10],
                   ['City','Total Drv','AA','GDR%','AR drv%','FR drv%','Online Hrs','Busy Hrs','Avg Income']):
    hcell(cur_r,col,txt)
cur_r+=1
for i,city in enumerate(CITIES):
    trow(cur_r); bg=BL if i%2==0 else WHITE
    td=drv1('E','$C$4',city); aa=drv1('L','$C$4',city)
    gdr=f"=IFERROR({drv1('F','$C$4',city)[1:]}/{td[1:]},0)"
    for val,fmt,col in zip([city,td,aa,gdr,drv_ar1('$C$4',city),drv_fr1('$C$4',city),
                            drv1('G','$C$4',city),drv1('H','$C$4',city),drv_income('$C$4',city)],
                            ['@','#,##0','#,##0','0.0%','0.0%','0.0%','#,##0.0','#,##0.0','#,##0'],
                            [2,3,4,5,6,7,8,9,10]):
        c=ws.cell(row=cur_r,column=col,value=val)
        c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()
    cur_r+=1
rh(cur_r,4); cur_r+=1

# Driver by Segment
r=cur_r; subsection(r,'Active Driver by Segment  (all cities combined)')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5,6,7,8,9,10],
                   ['Segment','Total Drv','AA','GDR%','AR drv%','FR drv%','Online Hrs','Busy Hrs','Avg Income']):
    hcell(cur_r,col,txt)
cur_r+=1
seg_colors_hex={'FT':GREEN,'NIM':BLUE,'NLM':'1D4ED8','PT':'7C3AED','Return':ORANGE}
for i,seg in enumerate(SEGS):
    trow(cur_r); bg=BL if i%2==0 else WHITE
    td=drv_seg1('E','$C$4',seg); aa=drv_seg1('L','$C$4',seg)
    gdr=f"=IFERROR({drv_seg1('F','$C$4',seg)[1:]}/{td[1:]},0)"
    for val,fmt,col in zip([seg,td,aa,gdr,drv_seg_ar('$C$4',seg),drv_seg_fr('$C$4',seg),
                            drv_seg1('G','$C$4',seg),drv_seg1('H','$C$4',seg),drv_income('$C$4',seg=seg)],
                            ['@','#,##0','#,##0','0.0%','0.0%','0.0%','#,##0.0','#,##0.0','#,##0'],
                            [2,3,4,5,6,7,8,9,10]):
        c=ws.cell(row=cur_r,column=col,value=val)
        c.font=fn(bold=False,color=seg_colors_hex.get(seg,DARK) if col==2 else DARK,size=9)
        c.fill=mf(bg); c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()
    cur_r+=1
rh(cur_r,4); cur_r+=1

# ─── SUSPENDED DRIVERS ────────────────────────────────────────────────────────
r=cur_r; section(r,'SUSPENDED DRIVERS  —  Latest Week  ($G$4)')
cur_r+=1

# By City
r=cur_r; subsection(r,'Total Suspended by City')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4],['City','Suspend Cases','% of Total']):
    hcell(cur_r,col,txt)
cur_r+=1
total_sus=f"SUMIF({SUS}!$B:$B,$G$4,{SUS}!$E:$E)"
for i,city in enumerate(CITIES):
    trow(cur_r); bg=BL if i%2==0 else WHITE
    tot=f"={sp1(SUS,'E','$G$4','C',city,SUS_N)}"
    pct=f"=IFERROR({tot[1:]}/({total_sus}),0)"
    for val,fmt,col in zip([city,tot,pct],['@','#,##0','0.0%'],[2,3,4]):
        c=ws.cell(row=cur_r,column=col,value=val)
        c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()
    cur_r+=1
rh(cur_r,4); cur_r+=1

# By Reason x City (pivot)
r=cur_r; subsection(r,'Suspended by Reason × City  (Suspend Flow)')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5,6],['Suspend Reason','SGN','HAN','EXP','Total']):
    hcell(cur_r,col,txt)
cur_r+=1
all_reasons = (sus_df[sus_df['period']==sus_df['period'].max()]
               .groupby('reason')['cases'].sum()
               .sort_values(ascending=False).index.tolist())
for i,rsn in enumerate(all_reasons[:15]):
    trow(cur_r); bg=BL if i%2==0 else WHITE
    rsn_s=rsn[:50]
    sgn=f"={sp2(SUS,'E','$G$4','C','SGN','D',rsn_s,SUS_N)}"
    han=f"={sp2(SUS,'E','$G$4','C','HAN','D',rsn_s,SUS_N)}"
    exp=f"={sp2(SUS,'E','$G$4','C','EXP','D',rsn_s,SUS_N)}"
    tot=f"={sp1(SUS,'E','$G$4','D',rsn_s,SUS_N)}"
    for val,fmt,col in zip([rsn_s,sgn,han,exp,tot],['@','#,##0','#,##0','#,##0','#,##0'],[2,3,4,5,6]):
        c=ws.cell(row=cur_r,column=col,value=val)
        c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()
    cur_r+=1
rh(cur_r,6); cur_r+=1

# ─── HOURLY PERFORMANCE ───────────────────────────────────────────────────────
r=cur_r; section(r,'HOURLY PERFORMANCE  —  Latest Week Aggregate  ($F$4)  (HOURLY DATA)')
cur_r+=1

# Overall by Hour
r=cur_r; subsection(r,'Overall — Requested / AR% / FR% / Surge Rate by Hour')
cur_r+=1; trow(cur_r,19)
for col,txt in zip([2,3,4,5,6,7,8],['Hour','Requested','Accept','AR%','Completed','FR%','Surge Rate']):
    hcell(cur_r,col,txt)
cur_r+=1
for hr in range(24):
    trow(cur_r); bg=BL if hr%2==0 else WHITE
    hr_ref=str(hr)
    req=f"={hrl('G','$F$4',hr_ref)}"
    acc=f"={hrl('H','$F$4',hr_ref)}"
    cmp=f"={hrl('I','$F$4',hr_ref)}"
    ar=hrl_ar('$F$4',hr_ref); fr=hrl_fr('$F$4',hr_ref)
    surge=hrl_surge('$F$4',hr_ref)
    lbl=f"{hr:02d}:00-{(hr+1)%24:02d}:00"
    for val,fmt,col in zip([lbl,req,acc,ar,cmp,fr,surge],
                            ['@','#,##0','#,##0','0.0%','#,##0','0.0%','0.0%'],[2,3,4,5,6,7,8]):
        c=ws.cell(row=cur_r,column=col,value=val)
        c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('center' if col>=3 else 'left'); c.border=mb()
    cur_r+=1
rh(cur_r,4); cur_r+=1

# Hourly by Region (wide: cols 2=Hour, then per region: Req, AR%, FR%)
r=cur_r; subsection(r,'By Region × Hour — Requested | AR% | FR%')
cur_r+=1; trow(cur_r,14)
ws.merge_cells(start_row=cur_r,start_column=2,end_row=cur_r,end_column=2)
ws.cell(row=cur_r,column=2,value='Hour').font=fn(bold=True,color=WHITE,size=8)
ws.cell(row=cur_r,column=2).fill=mf(BLUE); ws.cell(row=cur_r,column=2).alignment=al()
col_off=3
for reg in REGIONS:
    ws.merge_cells(start_row=cur_r,start_column=col_off,end_row=cur_r,end_column=col_off+2)
    c=ws.cell(row=cur_r,column=col_off,value=reg)
    c.font=fn(bold=True,color=WHITE,size=8); c.fill=mf(BLUE); c.alignment=al(); c.border=mb(BLUE)
    for cc,txt in zip([col_off,col_off+1,col_off+2],['Req','AR%','FR%']):
        ws.cell(row=cur_r+1,column=cc,value=txt).font=fn(bold=True,color=WHITE,size=8)
        ws.cell(row=cur_r+1,column=cc).fill=mf(BLUE)
        ws.cell(row=cur_r+1,column=cc).alignment=al()
        ws.cell(row=cur_r+1,column=cc).border=mb(BLUE)
    col_off+=3
ws.cell(row=cur_r,column=2).border=mb(BLUE)
cur_r+=2
for hr in range(24):
    trow(cur_r); bg=BL if hr%2==0 else WHITE
    hr_ref=str(hr)
    lbl=f"{hr:02d}:00"
    c=ws.cell(row=cur_r,column=2,value=lbl)
    c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg); c.alignment=al('center'); c.border=mb()
    col_off=3
    for reg in REGIONS:
        req=f"={hrl('G','$F$4',hr_ref,reg=reg)}"
        ar=hrl_ar('$F$4',hr_ref,reg=reg); fr=hrl_fr('$F$4',hr_ref,reg=reg)
        for val,fmt,col in zip([req,ar,fr],['#,##0','0.0%','0.0%'],[col_off,col_off+1,col_off+2]):
            c2=ws.cell(row=cur_r,column=col,value=val)
            c2.font=fn(bold=False,color=DARK,size=9); c2.fill=mf(bg)
            c2.number_format=fmt; c2.alignment=al('center'); c2.border=mb()
        col_off+=3
    cur_r+=1
rh(cur_r,4); cur_r+=1

# Hourly by Service (cols 2=Hour, then per service: Req, AR%)
r=cur_r; subsection(r,'By Service × Hour — Requested | AR%')
cur_r+=1; trow(cur_r,14)
ws.cell(row=cur_r,column=2,value='Hour').font=fn(bold=True,color=WHITE,size=8)
ws.cell(row=cur_r,column=2).fill=mf(BLUE); ws.cell(row=cur_r,column=2).alignment=al()
ws.cell(row=cur_r,column=2).border=mb(BLUE)
col_off=3
for svc_s in SERVICES:
    ws.merge_cells(start_row=cur_r,start_column=col_off,end_row=cur_r,end_column=col_off+1)
    c=ws.cell(row=cur_r,column=col_off,value=svc_s)
    c.font=fn(bold=True,color=WHITE,size=8); c.fill=mf(BLUE); c.alignment=al(); c.border=mb(BLUE)
    for cc,txt in zip([col_off,col_off+1],['Req','AR%']):
        ws.cell(row=cur_r+1,column=cc,value=txt).font=fn(bold=True,color=WHITE,size=8)
        ws.cell(row=cur_r+1,column=cc).fill=mf(BLUE)
        ws.cell(row=cur_r+1,column=cc).alignment=al()
        ws.cell(row=cur_r+1,column=cc).border=mb(BLUE)
    col_off+=2
cur_r+=2
for hr in range(24):
    trow(cur_r); bg=BL if hr%2==0 else WHITE
    hr_ref=str(hr)
    lbl=f"{hr:02d}:00"
    c=ws.cell(row=cur_r,column=2,value=lbl)
    c.font=fn(bold=False,color=DARK,size=9); c.fill=mf(bg); c.alignment=al('center'); c.border=mb()
    col_off=3
    for svc_s in SERVICES:
        req=f"={hrl('G','$F$4',hr_ref,svc=svc_s)}"
        ar=hrl_ar('$F$4',hr_ref,svc=svc_s)
        for val,fmt,col in zip([req,ar],['#,##0','0.0%'],[col_off,col_off+1]):
            c2=ws.cell(row=cur_r,column=col,value=val)
            c2.font=fn(bold=False,color=DARK,size=9); c2.fill=mf(bg)
            c2.number_format=fmt; c2.alignment=al('center'); c2.border=mb()
        col_off+=2
    cur_r+=1
rh(cur_r,6); cur_r+=1

# ─── DFD TREND ────────────────────────────────────────────────────────────────
r=cur_r; section(r,'DFD (NEW DRIVER) WEEKLY TREND  —  6 Weeks',GREEN)
cur_r+=1
r=cur_r; trow(r,19)
for col,txt in zip([2,3,4,5,6,7],['Week','Total DFD','Organic','Referral','Digital Ads','WoW %']):
    hcell(r,col,txt)
cur_r+=1
dfd_params=['$H$2','$I$2','$J$2','$K$2','$L$2','$M$2']
for i,wc in enumerate(dfd_params):
    trow(cur_r,17); bg=LGREEN if i==0 else (BL if i%2==0 else WHITE)
    nwc=dfd_params[i+1] if i<5 else None
    dfd_tot=f"=SUMIF({DFD}!$B:$B,{wc},{DFD}!$D:$D)"
    wow=f"=IFERROR({dfd_tot[1:]}/SUMIF({DFD}!$B:$B,{nwc},{DFD}!$D:$D)-1,\"—\")" if nwc else '"—"'
    lbl=f'=IFERROR(IF({wc}="","","W/E "&TEXT({wc},"DD/MM/YY")),"")'
    for val,fmt,col in zip([lbl,dfd_tot,
                            f"=SUMIF({DFD}!$B:$B,{wc},{DFD}!$E:$E)",
                            f"=SUMIF({DFD}!$B:$B,{wc},{DFD}!$F:$F)",
                            f"=SUMIF({DFD}!$B:$B,{wc},{DFD}!$G:$G)", wow],
                            ['@','#,##0','#,##0','#,##0','#,##0','0.0%'],[2,3,4,5,6,7]):
        c=ws.cell(row=cur_r,column=col,value=val)
        c.font=fn(bold=(i==0),color=DARK,size=9); c.fill=mf(bg)
        c.number_format=fmt; c.alignment=al('left' if col==2 else 'center'); c.border=mb()
    cur_r+=1

# ─── CONDITIONAL FORMATTING ───────────────────────────────────────────────────
from openpyxl.formatting.rule import ColorScaleRule
# (apply to key percentage columns if needed — kept minimal to avoid file bloat)

# ─── FREEZE + FINAL ───────────────────────────────────────────────────────────
ws.freeze_panes = 'B10'
ws.sheet_view.showGridLines = False

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Saved: {OUT}  ({os.path.getsize(OUT):,} bytes)  — Dashboard rows: ~{cur_r}")
